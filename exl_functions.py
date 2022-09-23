from openpyxl.reader.excel import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.chart import LineChart, Reference
from typing import Optional, List


def get_data(file_path: str, column: int) -> Optional[List[int]]:
    """Reads data from excel file and returns it as list"""
    try:
        worksheet = load_workbook(file_path).active

        result_values = list()
        # Getting column with power values
        for row in worksheet.iter_rows(min_row=2, min_col=column + 1, max_col=column + 1, values_only=True):
            result_values.append(row[0])

        return result_values
    except InvalidFileException:
        print("Invalid file!")
        return None


def write_data(file_path: str, column: int, data: list) -> None:
    """Writes data to existing excel file in the specified column"""
    try:
        workbook = load_workbook(file_path)
        worksheet = workbook.active
        # Filling cells with values
        for row_ind, value in enumerate(data, 2):
            temp_cell = worksheet.cell(row=row_ind, column=column + 1)
            temp_cell.value = value

        workbook.save(file_path)
    except InvalidFileException:
        print("Invalid file!")
        pass


def plot_results(file_path: str, max_col: int = 6, data_col: int = 2) -> None:
    """Plots prediction results in Excel file. Max column shows end of columns with predictions"""
    try:
        workbook = load_workbook(file_path)
        worksheet = workbook.active

        # Plotting base data
        data_chart = LineChart()

        # Chart configuration
        data_chart.title = "Энергопотребление за 24 часа"
        data_chart.x_axis.title = "Час"
        data_chart.y_axis.title = "Кол-во потреблённой энергии"
        data_chart.height = 10
        data_chart.width = 20

        # Configuring data
        data = Reference(worksheet, min_col=data_col, max_col=data_col, min_row=1, max_row=worksheet.max_row)
        x_label = Reference(worksheet, min_col=1, max_col=1, min_row=1, max_row=worksheet.max_row)

        # Setting data
        data_chart.add_data(data)
        data_chart.set_categories(x_label)

        worksheet.add_chart(data_chart, 'H2')

        # Plotting prediction results
        pred_chart = LineChart()

        # Chart configuration
        pred_chart.title = "Результаты работы алгоритмов"
        pred_chart.x_axis.title = "Час"
        pred_chart.y_axis.title = "Кол-во потреблённой энергии"
        pred_chart.height = 10
        pred_chart.width = 20

        # Configuring data
        data = Reference(worksheet, min_col=data_col + 1, max_col=max_col, min_row=1, max_row=worksheet.max_row)

        # Setting data
        pred_chart.add_data(data, titles_from_data=True)

        worksheet.add_chart(pred_chart, 'H22')

        workbook.save(file_path)
    except InvalidFileException:
        print("Invalid file!")
        pass
